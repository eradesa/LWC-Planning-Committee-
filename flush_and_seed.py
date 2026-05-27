#!/usr/bin/env python3
"""Flush all data and seed from sub-programs.xlsx.

Usage: DATABASE_URL="..." python3 flush_and_seed.py
"""
import os, sys, re
from datetime import date, timedelta

os.environ["DATABASE_URL"] = os.environ.get("DATABASE_URL") or "dbname=chms_dev"

import openpyxl
from models import get_conn, init_db, get_members, add_member
from models import get_program_categories, add_program_category
from models import add_sub_program, add_task, add_sub_program_member

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "sub-programs.xlsx")


def flush_all():
    conn = get_conn()
    tables = [
        "task_updates", "tasks", "sub_program_members",
        "sub_programs", "events", "program_categories",
        "members", "app_config", "users",
    ]
    for t in tables:
        conn.execute(f"TRUNCATE TABLE {t} RESTART IDENTITY CASCADE")
    conn.commit()
    conn.close()
    print("  Truncated all 9 tables")


def next_weekday(d, weekday):
    days_ahead = weekday - d.weekday()
    if days_ahead < 0:
        days_ahead += 7
    return d + timedelta(days=days_ahead)


SEED_MEMBERS = [
    "PASTOR DINUSHA", "PASTOR PRASAD", "PASTOR SHAMIKA",
    "PASTOR DINESH", "ROCHELLE", "SACHA", "ANITA",
    "SUREN", "YOHAN", "THERIKA", "RAJITH",
]


def seed_from_xlsx():
    today = date.today()
    next_sunday = next_weekday(today, 6)

    if today.month == 12:
        first_of_next_month = date(today.year + 1, 1, 1)
    else:
        first_of_next_month = date(today.year, today.month + 1, 1)

    q = (today.month - 1) // 3
    first_of_next_quarter = date(today.year, (q + 1) * 3 + 1, 1)

    for mn in SEED_MEMBERS:
        add_member(mn, "Member", "", "")
    all_members = get_members()
    print(f"  Seeded {len(all_members)} members")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    categories = {}
    subs = []
    current_sp = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        cat_name, flag, name, freq = [str(v).strip() if v else "" for v in row]

        if cat_name and flag == "Sub-program":
            categories[cat_name] = None

    cat_names_sorted = sorted(categories.keys())
    for cn in cat_names_sorted:
        add_program_category(cn, "", cat_names_sorted.index(cn) + 1)
        print(f"  Category: {cn}")

    cat_lookup = {c["name"]: c["id"] for c in get_program_categories()}
    member_lookup = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cat_name, flag, name, freq = [str(v).strip() if v else "" for v in row]

        if flag == "Sub-program" and name:
            freq_map = {
                "WEEKLY": "weekly", "BI-WEEKLY": "bi_weekly",
                "MONTHLY": "monthly", "QUARTERLY": "quarterly",
                "ANNUAL": "annual",
            }
            sp_freq = freq_map.get(freq.upper(), "none")

            if sp_freq == "weekly":
                due = next_sunday.isoformat()
            elif sp_freq == "bi_weekly":
                due = (today + timedelta(days=14)).isoformat()
            elif sp_freq == "monthly":
                due = first_of_next_month.isoformat()
            elif sp_freq == "quarterly":
                due = first_of_next_quarter.isoformat()
            elif sp_freq == "annual":
                try:
                    next_ann = date(today.year + 1, today.month, today.day)
                except ValueError:
                    next_ann = date(today.year + 1, today.month, 28)
                due = next_ann.isoformat()
            else:
                due = None

            cat_id = cat_lookup.get(cat_name)
            if not cat_id:
                print(f"  WARN: category '{cat_name}' not found, skipping")
                continue

            sp_id = add_sub_program(
                category_id=cat_id,
                title=name,
                description="",
                due_date=due,
                in_charge_id=None,
                recurring_type=sp_freq,
                add_to_calendar=1 if sp_freq != "none" else 0,
                type_flag="Program",
                notes="",
            )
            subs.append({"id": sp_id, "cat": cat_name, "sp": name, "freq": sp_freq})
            current_sp = sp_id
            print(f"  Sub-program: {name} ({sp_freq}) due {due}")

        elif flag == "Task" and name and current_sp is not None:
            task_due = due
            m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', name)
            if m:
                try:
                    td = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    task_due = td.isoformat()
                except ValueError:
                    pass
            add_task(
                sub_program_id=current_sp,
                title=name,
                due_date=task_due,
                assigned_to=None,
                priority="medium",
            )
            note = f" due {task_due}" if task_due else ""
            print(f"    Task: {name[:50]}{note}")

    wb.close()
    print(f"\nSeeded {len(cat_names_sorted)} categories, {len(subs)} sub-programs")


if __name__ == "__main__":
    print("Flushing all data...")
    flush_all()
    init_db()
    print("Seeding from sub-programs.xlsx...")
    seed_from_xlsx()
    print("Done.")
